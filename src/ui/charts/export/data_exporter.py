"""Structured data exporter for health visualizations."""

import pandas as pd
import json
import csv
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
from datetime import datetime
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .export_models import DataExportOptions, ExportResult

logger = logging.getLogger(__name__)


class StructuredDataExporter:
    """Exports visualization data in structured formats."""
    
    def __init__(self):
        """Initialize data exporter."""
        self.default_options = DataExportOptions()
        
    def export_dataframe(self, df: pd.DataFrame, format: str,
                        file_path: Union[str, Path],
                        options: Optional[DataExportOptions] = None) -> ExportResult:
        """Export pandas DataFrame in specified format."""
        options = options or self.default_options
        
        try:
            if format.lower() == 'csv':
                return self._export_csv(df, file_path, options)
            elif format.lower() == 'json':
                return self._export_json(df, file_path, options)
            elif format.lower() in ['excel', 'xlsx']:
                return self._export_excel(df, file_path, options)
            else:
                return ExportResult(
                    success=False,
                    error_message=f"Unsupported format: {format}"
                )
        except Exception as e:
            logger.error(f"Data export failed: {str(e)}")
            return ExportResult(success=False, error_message=str(e))
            
    def export_chart_data(self, chart: Any, format: str,
                         file_path: Union[str, Path],
                         options: Optional[DataExportOptions] = None) -> ExportResult:
        """Export chart data with metadata."""
        options = options or self.default_options
        
        try:
            # Extract data from chart
            data = self._extract_chart_data(chart, options)
            
            if isinstance(data, pd.DataFrame):
                return self.export_dataframe(data, format, file_path, options)
            elif isinstance(data, dict):
                return self._export_dict(data, format, file_path, options)
            else:
                return ExportResult(
                    success=False,
                    error_message="Unable to extract data from chart"
                )
        except Exception as e:
            logger.error(f"Chart data export failed: {str(e)}")
            return ExportResult(success=False, error_message=str(e))
            
    def export_dashboard_data(self, dashboard: Any, format: str,
                            output_dir: Union[str, Path],
                            options: Optional[DataExportOptions] = None) -> ExportResult:
        """Export all dashboard data."""
        options = options or self.default_options
        output_dir = Path(output_dir)
        
        try:
            # Create output directory if needed
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Collect all data
            all_data = {}
            metadata = {
                'export_date': datetime.now().strftime(options.date_format),
                'dashboard_title': getattr(dashboard, 'title', 'Health Dashboard'),
                'chart_count': len(dashboard.charts)
            }
            
            for chart_id, chart in dashboard.charts.items():
                chart_data = self._extract_chart_data(chart, options)
                if chart_data is not None:
                    all_data[chart_id] = chart_data
                    
            # Export based on format
            if format.lower() in ['excel', 'xlsx']:
                # Single Excel file with multiple sheets
                file_path = output_dir / f"dashboard_export.xlsx"
                return self._export_dashboard_excel(all_data, file_path, 
                                                  metadata, options)
            else:
                # Separate files for each chart
                for chart_id, data in all_data.items():
                    file_name = f"{chart_id}.{format.lower()}"
                    file_path = output_dir / file_name
                    
                    if isinstance(data, pd.DataFrame):
                        self.export_dataframe(data, format, file_path, options)
                    else:
                        self._export_dict(data, format, file_path, options)
                        
                # Also create metadata file
                meta_path = output_dir / f"metadata.json"
                with open(meta_path, 'w') as f:
                    json.dump(metadata, f, indent=2)
                    
                return ExportResult(
                    success=True,
                    file_path=output_dir,
                    metadata={'files_created': len(all_data) + 1}
                )
                
        except Exception as e:
            logger.error(f"Dashboard data export failed: {str(e)}")
            return ExportResult(success=False, error_message=str(e))
            
    def _extract_chart_data(self, chart: Any, 
                           options: DataExportOptions) -> Union[pd.DataFrame, Dict, None]:
        """Extract data from chart object."""
        # Try different methods to get data
        if hasattr(chart, 'get_export_data'):
            data = chart.get_export_data()
        elif hasattr(chart, 'get_data'):
            data = chart.get_data()
        elif hasattr(chart, 'data'):
            data = chart.data
        else:
            return None
            
        # Convert to DataFrame if needed
        if not isinstance(data, pd.DataFrame) and isinstance(data, (list, dict)):
            try:
                data = pd.DataFrame(data)
            except:
                return data  # Return as dict
                
        # Add metadata if requested
        if options.include_metadata and isinstance(data, pd.DataFrame):
            data = self._add_metadata_columns(data, chart, options)
            
        # Add calculations if available
        if options.include_calculations and hasattr(chart, 'get_calculations'):
            calcs = chart.get_calculations()
            if isinstance(data, pd.DataFrame):
                for calc_name, calc_value in calcs.items():
                    data[f'_calc_{calc_name}'] = calc_value
                    
        return data
        
    def _add_metadata_columns(self, df: pd.DataFrame, chart: Any,
                            options: DataExportOptions) -> pd.DataFrame:
        """Add metadata columns to DataFrame."""
        df = df.copy()
        
        # Add chart metadata
        df['_chart_type'] = chart.__class__.__name__
        df['_chart_title'] = getattr(chart, 'title', 'Unknown')
        df['_export_date'] = datetime.now().strftime(options.date_format)
        
        # Add data quality indicators if available
        if hasattr(chart, 'get_data_quality'):
            quality = chart.get_data_quality()
            df['_data_quality'] = quality.get('score', 'N/A')
            df['_completeness'] = quality.get('completeness', 'N/A')
            
        return df
        
    def _export_csv(self, df: pd.DataFrame, file_path: Union[str, Path],
                   options: DataExportOptions) -> ExportResult:
        """Export DataFrame as CSV."""
        try:
            df.to_csv(
                file_path,
                index=False,
                date_format=options.date_format,
                float_format=f'%.{options.decimal_places}f',
                na_rep=options.null_value
            )
            
            file_size = Path(file_path).stat().st_size
            
            return ExportResult(
                success=True,
                file_path=Path(file_path),
                file_size=file_size,
                metadata={'format': 'CSV', 'rows': len(df), 'columns': len(df.columns)}
            )
        except Exception as e:
            return ExportResult(success=False, error_message=str(e))
            
    def _export_json(self, df: pd.DataFrame, file_path: Union[str, Path],
                    options: DataExportOptions) -> ExportResult:
        """Export DataFrame as JSON."""
        try:
            # Convert DataFrame to dict
            data = df.to_dict(orient='records')
            
            # Add metadata wrapper if requested
            if options.include_metadata:
                export_data = {
                    'metadata': {
                        'export_date': datetime.now().isoformat(),
                        'record_count': len(data),
                        'columns': list(df.columns)
                    },
                    'data': data
                }
            else:
                export_data = data
                
            # Write JSON
            with open(file_path, 'w') as f:
                json.dump(export_data, f, indent=2, default=str)
                
            file_size = Path(file_path).stat().st_size
            
            return ExportResult(
                success=True,
                file_path=Path(file_path),
                file_size=file_size,
                metadata={'format': 'JSON', 'records': len(data)}
            )
        except Exception as e:
            return ExportResult(success=False, error_message=str(e))
            
    def _export_excel(self, df: pd.DataFrame, file_path: Union[str, Path],
                     options: DataExportOptions) -> ExportResult:
        """Export DataFrame as Excel with formatting."""
        try:
            # Create Excel writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write data
                sheet_name = 'Health Data'
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get worksheet for formatting
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting
                self._format_excel_sheet(worksheet, df, options)
                
            file_size = Path(file_path).stat().st_size
            
            return ExportResult(
                success=True,
                file_path=Path(file_path),
                file_size=file_size,
                metadata={'format': 'Excel', 'rows': len(df), 'columns': len(df.columns)}
            )
        except Exception as e:
            return ExportResult(success=False, error_message=str(e))
            
    def _export_dashboard_excel(self, all_data: Dict[str, Any], 
                              file_path: Union[str, Path],
                              metadata: Dict[str, Any],
                              options: DataExportOptions) -> ExportResult:
        """Export dashboard data as multi-sheet Excel."""
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Create summary sheet
                summary_df = pd.DataFrame([metadata])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary sheet
                summary_sheet = writer.sheets['Summary']
                self._format_summary_sheet(summary_sheet)
                
                # Add data sheets
                for chart_id, data in all_data.items():
                    # Truncate sheet name if too long
                    sheet_name = chart_id[:31] if len(chart_id) > 31 else chart_id
                    
                    if isinstance(data, pd.DataFrame):
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
                        worksheet = writer.sheets[sheet_name]
                        self._format_excel_sheet(worksheet, data, options)
                    elif isinstance(data, dict):
                        # Convert dict to DataFrame
                        dict_df = pd.DataFrame([data])
                        dict_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
            file_size = Path(file_path).stat().st_size
            
            return ExportResult(
                success=True,
                file_path=Path(file_path),
                file_size=file_size,
                metadata={
                    'format': 'Excel',
                    'sheets': len(all_data) + 1,
                    'total_rows': sum(len(d) if isinstance(d, pd.DataFrame) else 1 
                                    for d in all_data.values())
                }
            )
        except Exception as e:
            return ExportResult(success=False, error_message=str(e))
            
    def _format_excel_sheet(self, worksheet, df: pd.DataFrame, 
                          options: DataExportOptions):
        """Apply formatting to Excel worksheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF8C42", end_color="FF8C42", 
                                fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
        # Add alternating row colors
        for row in range(2, len(df) + 2):
            if row % 2 == 0:
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color="F5E6D3", 
                                          end_color="F5E6D3", 
                                          fill_type="solid")
                                          
        # Freeze header row
        worksheet.freeze_panes = 'A2'
        
    def _format_summary_sheet(self, worksheet):
        """Format summary sheet with WSJ styling."""
        # Title
        worksheet.merge_cells('A1:D1')
        title_cell = worksheet['A1']
        title_cell.value = "Dashboard Export Summary"
        title_cell.font = Font(size=16, bold=True, color="FF8C42")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add some spacing
        worksheet.row_dimensions[1].height = 30
        worksheet.row_dimensions[2].height = 20
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=3):
            for cell in row:
                if cell.column == 1:  # First column (keys)
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        # Auto-adjust columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    def _export_dict(self, data: Dict[str, Any], format: str,
                    file_path: Union[str, Path],
                    options: DataExportOptions) -> ExportResult:
        """Export dictionary data."""
        try:
            if format.lower() == 'json':
                with open(file_path, 'w') as f:
                    json.dump(data, f, indent=2, default=str)
            elif format.lower() == 'csv':
                # Convert dict to CSV (key-value pairs)
                with open(file_path, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Key', 'Value'])
                    for key, value in data.items():
                        writer.writerow([key, value])
            else:
                return ExportResult(
                    success=False,
                    error_message=f"Cannot export dict as {format}"
                )
                
            file_size = Path(file_path).stat().st_size
            
            return ExportResult(
                success=True,
                file_path=Path(file_path),
                file_size=file_size,
                metadata={'format': format.upper(), 'keys': len(data)}
            )
        except Exception as e:
            return ExportResult(success=False, error_message=str(e))
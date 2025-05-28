"""
Export & Reporting System for Apple Health Monitor.

This module provides comprehensive export and reporting functionality that allows users 
to generate professional health reports, export data in multiple formats, and create 
shareable insights for healthcare providers or personal tracking.
"""

import io
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass, field
import base64
from enum import Enum

import pandas as pd
import numpy as np
from jinja2 import Environment, FileSystemLoader, Template
from PyQt6.QtCore import QObject, pyqtSignal, QThread, QTimer
from PyQt6.QtWidgets import QProgressDialog
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as pdf_backend
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from ..ui.charts.wsj_style_manager import WSJStyleManager
from ..ui.charts.matplotlib_chart_factory import MatplotlibChartFactory
from .health_insights_engine import HealthInsightsEngine
from ..models import HealthData
from ..config import Config


class ExportFormat(Enum):
    """Supported export formats."""
    PDF = "pdf"
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    HTML = "html"
    PNG = "png"
    SVG = "svg"


class ReportTemplate(Enum):
    """Available report templates."""
    COMPREHENSIVE = "comprehensive"
    MEDICAL_SUMMARY = "medical_summary"
    PROGRESS_TRACKING = "progress_tracking"
    WEEKLY_SUMMARY = "weekly_summary"
    MONTHLY_SUMMARY = "monthly_summary"
    CUSTOM = "custom"


@dataclass
class ExportConfiguration:
    """Configuration for data export and report generation."""
    format: ExportFormat
    date_range: Tuple[datetime, datetime]
    metrics: List[str] = field(default_factory=list)
    include_charts: bool = True
    include_insights: bool = True
    include_raw_data: bool = False
    template: ReportTemplate = ReportTemplate.COMPREHENSIVE
    resolution: str = "high"  # "web", "print", "high"
    title: Optional[str] = None
    custom_sections: Dict[str, bool] = field(default_factory=dict)
    output_path: Optional[Path] = None
    
    def get_dpi(self) -> int:
        """Get DPI based on resolution setting."""
        dpi_map = {
            "web": 72,
            "print": 300,
            "high": 600
        }
        return dpi_map.get(self.resolution, 300)


class ExportProgressManager(QObject):
    """Manages export progress tracking and cancellation."""
    
    progress_updated = pyqtSignal(int, str)  # percent, message
    export_completed = pyqtSignal(str)  # result path
    export_failed = pyqtSignal(str)  # error message
    
    def __init__(self):
        super().__init__()
        self.is_cancelled = False
        self.current_progress = 0
        
    def start_export(self, export_type: str) -> 'ExportProgressManager':
        """Start tracking an export operation."""
        self.is_cancelled = False
        self.current_progress = 0
        self.progress_updated.emit(0, f"Starting {export_type} export...")
        return self
        
    def update(self, percent: int, message: str):
        """Update export progress."""
        if not self.is_cancelled:
            self.current_progress = percent
            self.progress_updated.emit(percent, message)
            
    def complete(self, result_path: str):
        """Mark export as complete."""
        if not self.is_cancelled:
            self.progress_updated.emit(100, "Export completed successfully")
            self.export_completed.emit(result_path)
            
    def error(self, error_message: str):
        """Mark export as failed."""
        self.export_failed.emit(error_message)
        
    def cancel(self):
        """Cancel the export operation."""
        self.is_cancelled = True
        self.progress_updated.emit(self.current_progress, "Export cancelled")


class WSJExportReportingSystem(QObject):
    """Professional export and reporting system following WSJ design principles."""
    
    def __init__(self, data_manager, viz_suite, insights_engine: HealthInsightsEngine, 
                 style_manager: WSJStyleManager):
        super().__init__()
        self.data_manager = data_manager
        self.viz_suite = viz_suite
        self.insights_engine = insights_engine
        self.style_manager = style_manager
        
        # Template system with WSJ styling
        self.template_env = self._setup_template_environment()
        self.template_registry = self._load_wsj_templates()
        
        # Export engines
        self.chart_factory = MatplotlibChartFactory(style_manager)
        self.export_progress = ExportProgressManager()
        
        # Export cache for performance
        self._export_cache = {}
        self._cache_timeout = 300  # 5 minutes
        
    def _setup_template_environment(self) -> Environment:
        """Set up Jinja2 template environment."""
        template_dir = Path(__file__).parent.parent / "templates" / "reports"
        template_dir.mkdir(parents=True, exist_ok=True)
        
        env = Environment(
            loader=FileSystemLoader(str(template_dir)),
            autoescape=True
        )
        
        # Add custom filters
        env.filters['format_date'] = lambda d: d.strftime('%B %d, %Y')
        env.filters['format_number'] = lambda n: f"{n:,.0f}"
        env.filters['format_percent'] = lambda n: f"{n:.1%}"
        
        return env
        
    def _load_wsj_templates(self) -> Dict[ReportTemplate, Template]:
        """Load WSJ-styled report templates."""
        templates = {}
        
        # Create inline templates for now
        templates[ReportTemplate.COMPREHENSIVE] = self._create_comprehensive_template()
        templates[ReportTemplate.MEDICAL_SUMMARY] = self._create_medical_template()
        templates[ReportTemplate.PROGRESS_TRACKING] = self._create_progress_template()
        templates[ReportTemplate.WEEKLY_SUMMARY] = self._create_weekly_template()
        templates[ReportTemplate.MONTHLY_SUMMARY] = self._create_monthly_template()
        
        return templates
        
    def generate_report(self, config: ExportConfiguration) -> Union[bytes, str, Dict]:
        """Generate report based on configuration."""
        if config.format == ExportFormat.PDF:
            return self.generate_wsj_pdf_report(config)
        elif config.format == ExportFormat.CSV:
            return self.export_data_csv(config)
        elif config.format == ExportFormat.EXCEL:
            return self.export_data_excel(config)
        elif config.format == ExportFormat.JSON:
            return self.export_data_json(config)
        elif config.format == ExportFormat.HTML:
            return self.generate_html_summary(config)
        elif config.format in [ExportFormat.PNG, ExportFormat.SVG]:
            return self.export_charts(config)
        else:
            raise ValueError(f"Unsupported export format: {config.format}")
            
    def generate_wsj_pdf_report(self, config: ExportConfiguration) -> bytes:
        """Generate professional PDF report with WSJ styling."""
        progress_tracker = self.export_progress.start_export('PDF report')
        
        try:
            # Gather data
            progress_tracker.update(10, "Gathering health data...")
            report_data = self._gather_report_data(config)
            
            # Generate insights
            progress_tracker.update(30, "Analyzing patterns and trends...")
            insights = self.insights_engine.generate_prioritized_insights(
                report_data['metrics'],
                max_insights=10
            )
            
            # Create charts
            progress_tracker.update(50, "Creating visualizations...")
            charts = self._generate_report_charts(report_data, config)
            
            # Create PDF
            progress_tracker.update(70, "Generating PDF document...")
            pdf_buffer = io.BytesIO()
            
            # Create document
            doc = SimpleDocTemplate(
                pdf_buffer,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18,
            )
            
            # Build content
            story = []
            styles = self._get_pdf_styles()
            
            # Title page
            story.append(Paragraph(config.title or "Health Analysis Report", styles['Title']))
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph(
                f"{config.date_range[0].strftime('%B %d, %Y')} - {config.date_range[1].strftime('%B %d, %Y')}",
                styles['Subtitle']
            ))
            story.append(Spacer(1, 0.5*inch))
            
            # Executive summary
            if insights:
                story.append(Paragraph("Executive Summary", styles['Heading1']))
                story.append(Spacer(1, 0.2*inch))
                
                # Key insight
                key_insight = insights[0]
                story.append(Paragraph(f"<b>Key Finding:</b> {key_insight.title}", styles['BodyText']))
                story.append(Paragraph(key_insight.summary, styles['BodyText']))
                story.append(Spacer(1, 0.3*inch))
                
                # Additional insights
                if len(insights) > 1:
                    story.append(Paragraph("<b>Additional Insights:</b>", styles['BodyText']))
                    for insight in insights[1:4]:  # Top 3 additional
                        story.append(Paragraph(f"• {insight.title}", styles['BodyText']))
                    story.append(Spacer(1, 0.3*inch))
            
            story.append(PageBreak())
            
            # Metrics overview
            story.append(Paragraph("Health Metrics Overview", styles['Heading1']))
            story.append(Spacer(1, 0.3*inch))
            
            # Add summary statistics table
            if report_data['summary_stats']:
                table_data = [['Metric', 'Average', 'Min', 'Max', 'Trend']]
                for metric, stats in report_data['summary_stats'].items():
                    table_data.append([
                        metric.replace('_', ' ').title(),
                        f"{stats['mean']:.1f}",
                        f"{stats['min']:.1f}",
                        f"{stats['max']:.1f}",
                        stats.get('trend', 'Stable')
                    ])
                
                table = Table(table_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1.5*inch])
                table.setStyle(self._get_table_style())
                story.append(table)
                story.append(Spacer(1, 0.5*inch))
            
            # Add charts
            for chart in charts:
                if chart['image']:
                    img = Image(io.BytesIO(chart['image']), width=6*inch, height=4*inch)
                    story.append(KeepTogether([
                        Paragraph(chart['title'], styles['Heading2']),
                        Spacer(1, 0.1*inch),
                        img,
                        Spacer(1, 0.1*inch),
                        Paragraph(chart['description'], styles['Caption']),
                        Spacer(1, 0.3*inch)
                    ]))
            
            story.append(PageBreak())
            
            # Detailed insights
            story.append(Paragraph("Detailed Analysis & Recommendations", styles['Heading1']))
            story.append(Spacer(1, 0.3*inch))
            
            for i, insight in enumerate(insights):
                story.append(Paragraph(f"{i+1}. {insight.title}", styles['Heading2']))
                story.append(Paragraph(f"<b>Summary:</b> {insight.summary}", styles['BodyText']))
                story.append(Paragraph(f"<b>Recommendation:</b> {insight.recommendation}", styles['BodyText']))
                
                if insight.evidence_level != 'pattern_based':
                    story.append(Paragraph(
                        f"<i>Evidence Level: {insight.evidence_level.replace('_', ' ').title()}</i>",
                        styles['Caption']
                    ))
                
                story.append(Spacer(1, 0.3*inch))
            
            # Footer
            story.append(PageBreak())
            story.append(Paragraph("Disclaimer", styles['Heading2']))
            story.append(Paragraph(
                "This report is for informational purposes only and does not constitute medical advice. "
                "Please consult with healthcare professionals for medical concerns.",
                styles['Disclaimer']
            ))
            
            # Build PDF
            doc.build(story)
            pdf_bytes = pdf_buffer.getvalue()
            
            progress_tracker.update(100, "PDF report generated successfully")
            progress_tracker.complete("report.pdf")
            
            return pdf_bytes
            
        except Exception as e:
            progress_tracker.error(f"Report generation failed: {str(e)}")
            raise
            
    def export_data_csv(self, config: ExportConfiguration) -> bytes:
        """Export data as CSV with proper formatting."""
        progress_tracker = self.export_progress.start_export('CSV export')
        
        try:
            progress_tracker.update(20, "Gathering data...")
            report_data = self._gather_report_data(config)
            
            progress_tracker.update(50, "Formatting CSV data...")
            output = io.StringIO()
            
            # Write metadata
            writer = csv.writer(output)
            writer.writerow(['Apple Health Monitor Export'])
            writer.writerow(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow(['Date Range:', f"{config.date_range[0]} to {config.date_range[1]}"])
            writer.writerow([])  # Empty row
            
            # Write data for each metric
            for metric in config.metrics:
                writer.writerow([f'Metric: {metric.replace("_", " ").title()}'])
                
                # Get data for metric
                df = report_data['raw_data'].get(metric, pd.DataFrame())
                if not df.empty:
                    # Write headers
                    writer.writerow(['Date', 'Value', 'Unit'])
                    
                    # Write data
                    for idx, row in df.iterrows():
                        writer.writerow([
                            idx.strftime('%Y-%m-%d %H:%M:%S'),
                            row.get('value', ''),
                            row.get('unit', '')
                        ])
                else:
                    writer.writerow(['No data available'])
                
                writer.writerow([])  # Empty row between metrics
            
            # Add summary statistics
            if config.include_insights and report_data['summary_stats']:
                writer.writerow(['Summary Statistics'])
                writer.writerow(['Metric', 'Count', 'Mean', 'Min', 'Max', 'Std Dev'])
                
                for metric, stats in report_data['summary_stats'].items():
                    writer.writerow([
                        metric.replace('_', ' ').title(),
                        stats.get('count', 0),
                        f"{stats.get('mean', 0):.2f}",
                        f"{stats.get('min', 0):.2f}",
                        f"{stats.get('max', 0):.2f}",
                        f"{stats.get('std', 0):.2f}"
                    ])
            
            progress_tracker.update(100, "CSV export completed")
            csv_bytes = output.getvalue().encode('utf-8')
            progress_tracker.complete("export.csv")
            
            return csv_bytes
            
        except Exception as e:
            progress_tracker.error(f"CSV export failed: {str(e)}")
            raise
            
    def export_data_excel(self, config: ExportConfiguration) -> bytes:
        """Export data as Excel with formatting and multiple sheets."""
        progress_tracker = self.export_progress.start_export('Excel export')
        
        try:
            progress_tracker.update(10, "Creating workbook...")
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            date_font = Font(italic=True)
            number_alignment = Alignment(horizontal="right")
            
            # Summary sheet
            progress_tracker.update(20, "Creating summary sheet...")
            summary_sheet = wb.create_sheet("Summary")
            
            # Title
            summary_sheet['A1'] = "Apple Health Monitor Export"
            summary_sheet['A1'].font = Font(bold=True, size=16)
            summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            summary_sheet['A3'] = f"Date Range: {config.date_range[0]} to {config.date_range[1]}"
            
            # Gather report data
            progress_tracker.update(30, "Gathering data...")
            report_data = self._gather_report_data(config)
            
            # Summary statistics
            if report_data['summary_stats']:
                row = 6
                summary_sheet[f'A{row}'] = "Summary Statistics"
                summary_sheet[f'A{row}'].font = Font(bold=True, size=12)
                
                row += 2
                headers = ['Metric', 'Count', 'Average', 'Min', 'Max', 'Std Dev']
                for col, header in enumerate(headers, 1):
                    cell = summary_sheet.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                row += 1
                for metric, stats in report_data['summary_stats'].items():
                    summary_sheet.cell(row=row, column=1, value=metric.replace('_', ' ').title())
                    summary_sheet.cell(row=row, column=2, value=stats.get('count', 0))
                    summary_sheet.cell(row=row, column=3, value=round(stats.get('mean', 0), 2))
                    summary_sheet.cell(row=row, column=4, value=round(stats.get('min', 0), 2))
                    summary_sheet.cell(row=row, column=5, value=round(stats.get('max', 0), 2))
                    summary_sheet.cell(row=row, column=6, value=round(stats.get('std', 0), 2))
                    row += 1
                
                # Adjust column widths
                for col in range(1, 7):
                    summary_sheet.column_dimensions[get_column_letter(col)].width = 15
            
            # Create sheets for each metric
            progress_tracker.update(50, "Creating metric sheets...")
            for i, metric in enumerate(config.metrics):
                sheet_name = metric.replace('_', ' ').title()[:31]  # Excel limit
                ws = wb.create_sheet(sheet_name)
                
                # Headers
                headers = ['Date', 'Time', 'Value', 'Unit']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Data
                df = report_data['raw_data'].get(metric, pd.DataFrame())
                if not df.empty:
                    row = 2
                    for idx, data_row in df.iterrows():
                        ws.cell(row=row, column=1, value=idx.date())
                        ws.cell(row=row, column=2, value=idx.time())
                        ws.cell(row=row, column=3, value=data_row.get('value', 0))
                        ws.cell(row=row, column=4, value=data_row.get('unit', ''))
                        row += 1
                    
                    # Format columns
                    ws.column_dimensions['A'].width = 12
                    ws.column_dimensions['B'].width = 10
                    ws.column_dimensions['C'].width = 10
                    ws.column_dimensions['D'].width = 10
                
                progress_tracker.update(50 + (40 * (i + 1) // len(config.metrics)), 
                                      f"Processing {sheet_name}...")
            
            # Insights sheet
            if config.include_insights:
                progress_tracker.update(90, "Adding insights...")
                insights_sheet = wb.create_sheet("Insights")
                
                insights = self.insights_engine.generate_prioritized_insights(
                    report_data['metrics'], 
                    max_insights=10
                )
                
                row = 1
                insights_sheet['A1'] = "Health Insights & Recommendations"
                insights_sheet['A1'].font = Font(bold=True, size=14)
                
                row = 3
                for i, insight in enumerate(insights, 1):
                    insights_sheet[f'A{row}'] = f"{i}. {insight.title}"
                    insights_sheet[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    insights_sheet[f'A{row}'] = f"Summary: {insight.summary}"
                    row += 1
                    
                    insights_sheet[f'A{row}'] = f"Recommendation: {insight.recommendation}"
                    row += 1
                    
                    insights_sheet[f'A{row}'] = f"Evidence Level: {insight.evidence_level}"
                    insights_sheet[f'A{row}'].font = Font(italic=True)
                    row += 2
                
                # Adjust column width
                insights_sheet.column_dimensions['A'].width = 100
            
            # Save to bytes
            progress_tracker.update(95, "Finalizing workbook...")
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_bytes = excel_buffer.getvalue()
            
            progress_tracker.update(100, "Excel export completed")
            progress_tracker.complete("export.xlsx")
            
            return excel_bytes
            
        except Exception as e:
            progress_tracker.error(f"Excel export failed: {str(e)}")
            raise
            
    def export_data_json(self, config: ExportConfiguration) -> bytes:
        """Export data as JSON with metadata."""
        progress_tracker = self.export_progress.start_export('JSON export')
        
        try:
            progress_tracker.update(20, "Gathering data...")
            report_data = self._gather_report_data(config)
            
            progress_tracker.update(50, "Formatting JSON data...")
            
            # Build JSON structure
            export_data = {
                'metadata': {
                    'export_date': datetime.now().isoformat(),
                    'date_range': {
                        'start': config.date_range[0].isoformat(),
                        'end': config.date_range[1].isoformat()
                    },
                    'version': '1.0',
                    'metrics': config.metrics
                },
                'summary_statistics': report_data['summary_stats'],
                'data': {}
            }
            
            # Add raw data for each metric
            for metric in config.metrics:
                df = report_data['raw_data'].get(metric, pd.DataFrame())
                if not df.empty:
                    metric_data = []
                    for idx, row in df.iterrows():
                        metric_data.append({
                            'timestamp': idx.isoformat(),
                            'value': float(row.get('value', 0)),
                            'unit': row.get('unit', '')
                        })
                    export_data['data'][metric] = metric_data
            
            # Add insights if requested
            if config.include_insights:
                insights = self.insights_engine.generate_prioritized_insights(
                    report_data['metrics'],
                    max_insights=10
                )
                export_data['insights'] = [
                    {
                        'title': insight.title,
                        'summary': insight.summary,
                        'recommendation': insight.recommendation,
                        'evidence_level': insight.evidence_level,
                        'confidence_score': insight.confidence_score,
                        'impact_score': insight.impact_score
                    }
                    for insight in insights
                ]
            
            progress_tracker.update(90, "Encoding JSON...")
            json_bytes = json.dumps(export_data, indent=2).encode('utf-8')
            
            progress_tracker.update(100, "JSON export completed")
            progress_tracker.complete("export.json")
            
            return json_bytes
            
        except Exception as e:
            progress_tracker.error(f"JSON export failed: {str(e)}")
            raise
            
    def generate_html_summary(self, config: ExportConfiguration) -> str:
        """Generate email-friendly HTML summary."""
        progress_tracker = self.export_progress.start_export('HTML summary')
        
        try:
            progress_tracker.update(20, "Gathering data...")
            report_data = self._gather_report_data(config)
            
            progress_tracker.update(40, "Generating insights...")
            insights = self.insights_engine.generate_prioritized_insights(
                report_data['metrics'],
                max_insights=5  # Limit for email
            )
            
            progress_tracker.update(60, "Creating visualizations...")
            # Generate small charts for email
            charts = []
            if config.include_charts:
                for metric in config.metrics[:3]:  # Limit charts for email
                    chart_data = self._generate_email_chart(metric, report_data)
                    if chart_data:
                        charts.append(chart_data)
            
            progress_tracker.update(80, "Rendering HTML...")
            
            # Use template to generate HTML
            template_data = {
                'title': config.title or "Health Summary",
                'date_range': config.date_range,
                'summary_stats': report_data['summary_stats'],
                'insights': insights,
                'charts': charts,
                'generated_date': datetime.now(),
                'style': self.style_manager.get_export_styles()
            }
            
            html_content = self._render_html_template(template_data)
            
            progress_tracker.update(100, "HTML summary completed")
            progress_tracker.complete("summary.html")
            
            return html_content
            
        except Exception as e:
            progress_tracker.error(f"HTML generation failed: {str(e)}")
            raise
            
    def export_charts(self, config: ExportConfiguration) -> Dict[str, bytes]:
        """Export charts as high-resolution images."""
        progress_tracker = self.export_progress.start_export('Chart export')
        
        try:
            progress_tracker.update(10, "Gathering data...")
            report_data = self._gather_report_data(config)
            
            charts = {}
            total_metrics = len(config.metrics)
            
            for i, metric in enumerate(config.metrics):
                progress_tracker.update(
                    10 + (80 * i // total_metrics),
                    f"Generating chart for {metric}..."
                )
                
                # Generate chart using matplotlib factory
                chart_bytes = self._generate_metric_chart(
                    metric, 
                    report_data,
                    format=config.format.value,
                    dpi=config.get_dpi()
                )
                
                if chart_bytes:
                    charts[metric] = chart_bytes
            
            progress_tracker.update(100, "Chart export completed")
            progress_tracker.complete(f"charts.{config.format.value}")
            
            return charts
            
        except Exception as e:
            progress_tracker.error(f"Chart export failed: {str(e)}")
            raise
            
    def create_backup(self, include_settings: bool = True) -> bytes:
        """Create complete data backup with integrity verification."""
        progress_tracker = self.export_progress.start_export('Backup')
        
        try:
            import zipfile
            import hashlib
            
            progress_tracker.update(10, "Preparing backup...")
            
            backup_buffer = io.BytesIO()
            
            with zipfile.ZipFile(backup_buffer, 'w', zipfile.ZIP_DEFLATED) as backup_zip:
                # Export all data as JSON
                progress_tracker.update(30, "Exporting health data...")
                
                # Get all available metrics
                all_metrics = self.data_manager.get_available_metrics()
                
                # Export configuration for full backup
                backup_config = ExportConfiguration(
                    format=ExportFormat.JSON,
                    date_range=(
                        datetime(2000, 1, 1),  # Far past
                        datetime.now()
                    ),
                    metrics=all_metrics,
                    include_charts=False,
                    include_insights=True,
                    include_raw_data=True
                )
                
                # Generate JSON data
                json_data = self.export_data_json(backup_config)
                backup_zip.writestr('health_data.json', json_data)
                
                # Add metadata
                metadata = {
                    'backup_date': datetime.now().isoformat(),
                    'version': '1.0',
                    'metrics_count': len(all_metrics),
                    'data_hash': hashlib.sha256(json_data).hexdigest()
                }
                backup_zip.writestr('metadata.json', json.dumps(metadata, indent=2))
                
                # Include settings if requested
                if include_settings:
                    progress_tracker.update(60, "Including settings...")
                    settings_data = self._export_settings()
                    backup_zip.writestr('settings.json', json.dumps(settings_data, indent=2))
                
                # Add verification file
                progress_tracker.update(80, "Creating verification data...")
                verification_data = {
                    'files': ['health_data.json', 'metadata.json'],
                    'checksums': {
                        'health_data.json': hashlib.sha256(json_data).hexdigest()
                    }
                }
                
                if include_settings:
                    verification_data['files'].append('settings.json')
                    verification_data['checksums']['settings.json'] = hashlib.sha256(
                        json.dumps(settings_data).encode()
                    ).hexdigest()
                
                backup_zip.writestr('verification.json', json.dumps(verification_data, indent=2))
            
            progress_tracker.update(100, "Backup completed")
            backup_bytes = backup_buffer.getvalue()
            progress_tracker.complete("backup.zip")
            
            return backup_bytes
            
        except Exception as e:
            progress_tracker.error(f"Backup failed: {str(e)}")
            raise
            
    def generate_shareable_insight(self, insight_id: str, format: str = "png") -> bytes:
        """Create shareable insight card image."""
        try:
            # Get insight data
            insight = self.insights_engine.get_insight_by_id(insight_id)
            if not insight:
                raise ValueError(f"Insight {insight_id} not found")
            
            # Create styled insight card
            fig, ax = plt.subplots(1, 1, figsize=(8, 6))
            ax.axis('off')
            
            # Apply WSJ styling
            style = self.style_manager.get_export_styles()
            fig.patch.set_facecolor(style['colors']['background'])
            
            # Title
            ax.text(0.5, 0.9, insight.title, 
                   fontsize=20, fontweight='bold',
                   ha='center', va='top',
                   color=style['colors']['primary'],
                   fontfamily=style['fonts']['headline'])
            
            # Summary
            ax.text(0.5, 0.7, insight.summary,
                   fontsize=14, ha='center', va='center',
                   wrap=True, multialignment='center',
                   color=style['colors']['text_primary'],
                   fontfamily=style['fonts']['body'])
            
            # Recommendation
            ax.text(0.5, 0.4, f"Recommendation: {insight.recommendation}",
                   fontsize=12, ha='center', va='center',
                   wrap=True, multialignment='center',
                   style='italic',
                   color=style['colors']['text_secondary'],
                   fontfamily=style['fonts']['body'])
            
            # Evidence level
            ax.text(0.5, 0.2, f"Evidence: {insight.evidence_level.replace('_', ' ').title()}",
                   fontsize=10, ha='center', va='center',
                   color=style['colors']['text_muted'],
                   fontfamily=style['fonts']['body'])
            
            # Footer
            ax.text(0.5, 0.05, "Generated by Apple Health Monitor",
                   fontsize=8, ha='center', va='bottom',
                   color=style['colors']['text_muted'],
                   fontfamily=style['fonts']['body'])
            
            # Save to bytes
            buffer = io.BytesIO()
            fig.savefig(buffer, format=format, dpi=300, bbox_inches='tight',
                       facecolor=style['colors']['background'])
            plt.close(fig)
            
            return buffer.getvalue()
            
        except Exception as e:
            raise RuntimeError(f"Failed to generate shareable insight: {str(e)}")
            
    # Helper methods
    
    def _gather_report_data(self, config: ExportConfiguration) -> Dict[str, Any]:
        """Gather all data needed for report generation."""
        report_data = {
            'metrics': {},
            'raw_data': {},
            'summary_stats': {},
            'date_range': config.date_range
        }
        
        for metric in config.metrics:
            # Get data from data manager
            df = self.data_manager.get_metric_data(
                metric, 
                config.date_range[0], 
                config.date_range[1]
            )
            
            if df is not None and not df.empty:
                report_data['raw_data'][metric] = df
                report_data['metrics'][metric] = df
                
                # Calculate summary statistics
                if 'value' in df.columns:
                    report_data['summary_stats'][metric] = {
                        'count': len(df),
                        'mean': df['value'].mean(),
                        'min': df['value'].min(),
                        'max': df['value'].max(),
                        'std': df['value'].std(),
                        'trend': self._calculate_trend(df['value'])
                    }
        
        return report_data
        
    def _generate_report_charts(self, report_data: Dict, config: ExportConfiguration) -> List[Dict]:
        """Generate charts for the report."""
        charts = []
        
        # Overview chart - combined metrics
        if len(config.metrics) > 1:
            overview_chart = self._generate_overview_chart(report_data, config)
            if overview_chart:
                charts.append(overview_chart)
        
        # Individual metric charts
        for metric in config.metrics[:5]:  # Limit to 5 charts
            metric_chart = self._generate_metric_detail_chart(metric, report_data, config)
            if metric_chart:
                charts.append(metric_chart)
        
        return charts
        
    def _generate_overview_chart(self, report_data: Dict, config: ExportConfiguration) -> Optional[Dict]:
        """Generate overview chart combining multiple metrics."""
        try:
            # Use matplotlib factory to create multi-metric chart
            chart_config = {
                'title': 'Health Metrics Overview',
                'style': 'wsj_publication',
                'size': (10, 6),
                'dpi': config.get_dpi()
            }
            
            # Prepare data for visualization
            metrics_data = []
            for metric, df in report_data['raw_data'].items():
                if not df.empty and 'value' in df.columns:
                    # Normalize values for comparison
                    normalized = (df['value'] - df['value'].mean()) / df['value'].std()
                    metrics_data.append({
                        'name': metric.replace('_', ' ').title(),
                        'data': normalized.resample('D').mean()  # Daily averages
                    })
            
            if not metrics_data:
                return None
            
            # Create chart
            fig, ax = plt.subplots(figsize=chart_config['size'])
            
            # Apply WSJ styling
            self.style_manager.apply_wsj_style(fig, ax)
            
            # Plot each metric
            for metric_info in metrics_data[:4]:  # Limit to 4 metrics
                ax.plot(metric_info['data'].index, metric_info['data'].values,
                       label=metric_info['name'], linewidth=2)
            
            ax.set_title(chart_config['title'])
            ax.set_xlabel('Date')
            ax.set_ylabel('Normalized Value')
            ax.legend(loc='best')
            
            # Convert to bytes
            buffer = io.BytesIO()
            fig.savefig(buffer, format='png', dpi=chart_config['dpi'], bbox_inches='tight')
            plt.close(fig)
            
            return {
                'title': chart_config['title'],
                'description': 'Normalized comparison of health metrics over time',
                'image': buffer.getvalue()
            }
            
        except Exception as e:
            print(f"Error generating overview chart: {e}")
            return None
            
    def _generate_metric_detail_chart(self, metric: str, report_data: Dict, 
                                    config: ExportConfiguration) -> Optional[Dict]:
        """Generate detailed chart for a single metric."""
        try:
            df = report_data['raw_data'].get(metric)
            if df is None or df.empty:
                return None
            
            # Create chart using matplotlib factory
            chart = self.chart_factory.create_chart(
                chart_type='line',
                data=df,
                config={
                    'title': f"{metric.replace('_', ' ').title()} Over Time",
                    'x_label': 'Date',
                    'y_label': df.get('unit', ['Value'])[0] if 'unit' in df else 'Value',
                    'style': 'wsj_publication',
                    'size': (8, 5),
                    'dpi': config.get_dpi()
                }
            )
            
            # Convert to bytes
            buffer = io.BytesIO()
            chart.figure.savefig(buffer, format='png', dpi=config.get_dpi(), bbox_inches='tight')
            plt.close(chart.figure)
            
            return {
                'title': f"{metric.replace('_', ' ').title()} Trends",
                'description': f"Daily {metric.replace('_', ' ')} measurements over the selected period",
                'image': buffer.getvalue()
            }
            
        except Exception as e:
            print(f"Error generating chart for {metric}: {e}")
            return None
            
    def _generate_metric_chart(self, metric: str, report_data: Dict, 
                             format: str = "png", dpi: int = 300) -> Optional[bytes]:
        """Generate a single metric chart for export."""
        try:
            df = report_data['raw_data'].get(metric)
            if df is None or df.empty:
                return None
            
            # Create figure
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Apply WSJ styling
            self.style_manager.apply_wsj_style(fig, ax)
            
            # Plot data
            if 'value' in df.columns:
                ax.plot(df.index, df['value'], linewidth=2, color=self.style_manager.colors['primary'])
                ax.fill_between(df.index, df['value'], alpha=0.1, color=self.style_manager.colors['primary'])
            
            # Labels
            ax.set_title(f"{metric.replace('_', ' ').title()}")
            ax.set_xlabel('Date')
            ax.set_ylabel(df.get('unit', ['Value'])[0] if 'unit' in df else 'Value')
            
            # Format
            buffer = io.BytesIO()
            fig.savefig(buffer, format=format, dpi=dpi, bbox_inches='tight')
            plt.close(fig)
            
            return buffer.getvalue()
            
        except Exception as e:
            print(f"Error generating chart for {metric}: {e}")
            return None
            
    def _generate_email_chart(self, metric: str, report_data: Dict) -> Optional[Dict]:
        """Generate small chart suitable for email embedding."""
        try:
            df = report_data['raw_data'].get(metric)
            if df is None or df.empty:
                return None
            
            # Create small figure
            fig, ax = plt.subplots(figsize=(4, 3))
            
            # Simple styling for email
            ax.plot(df.index, df['value'], linewidth=1.5)
            ax.set_title(metric.replace('_', ' ').title(), fontsize=10)
            ax.tick_params(labelsize=8)
            
            # Convert to base64 for email embedding
            buffer = io.BytesIO()
            fig.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close(fig)
            
            image_base64 = base64.b64encode(buffer.getvalue()).decode()
            
            return {
                'metric': metric,
                'title': metric.replace('_', ' ').title(),
                'data': image_base64
            }
            
        except Exception as e:
            print(f"Error generating email chart for {metric}: {e}")
            return None
            
    def _calculate_trend(self, series: pd.Series) -> str:
        """Calculate trend direction for a data series."""
        if len(series) < 2:
            return "Insufficient data"
        
        # Simple linear regression
        x = np.arange(len(series))
        y = series.values
        
        # Handle NaN values
        mask = ~np.isnan(y)
        if mask.sum() < 2:
            return "Insufficient data"
        
        x = x[mask]
        y = y[mask]
        
        # Calculate slope
        slope = np.polyfit(x, y, 1)[0]
        
        # Determine trend
        if abs(slope) < 0.01:
            return "Stable"
        elif slope > 0:
            return "Increasing"
        else:
            return "Decreasing"
            
    def _export_settings(self) -> Dict:
        """Export current application settings."""
        # This would be implemented based on your settings system
        return {
            'export_preferences': {
                'default_format': 'pdf',
                'default_resolution': 'high',
                'include_insights': True,
                'include_charts': True
            },
            'display_preferences': {
                'theme': 'wsj',
                'colors': self.style_manager.colors
            }
        }
        
    def _get_pdf_styles(self) -> Dict:
        """Get PDF paragraph styles matching WSJ design."""
        styles = getSampleStyleSheet()
        
        # Title style
        styles.add(ParagraphStyle(
            name='Title',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C2C2C'),
            spaceAfter=12,
            alignment=TA_CENTER
        ))
        
        # Subtitle style
        styles.add(ParagraphStyle(
            name='Subtitle',
            parent=styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor('#6B6B6B'),
            alignment=TA_CENTER
        ))
        
        # Body text style
        styles.add(ParagraphStyle(
            name='BodyText',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#2C2C2C'),
            spaceAfter=8,
            leading=14
        ))
        
        # Caption style
        styles.add(ParagraphStyle(
            name='Caption',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6B6B6B'),
            alignment=TA_CENTER,
            italic=True
        ))
        
        # Disclaimer style
        styles.add(ParagraphStyle(
            name='Disclaimer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6B6B6B'),
            italic=True
        ))
        
        return styles
        
    def _get_table_style(self) -> TableStyle:
        """Get table style for PDF reports."""
        return TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF8C42')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2C2C2C')),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E8DCC8')),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#8B7355')),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAF8F5')])
        ])
        
    def _render_html_template(self, data: Dict) -> str:
        """Render HTML template with data."""
        # Create a simple HTML template inline
        template = Template("""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6;
            color: #2C2C2C;
            background-color: #FFFFFF;
            margin: 0;
            padding: 20px;
        }
        
        .container {
            max-width: 600px;
            margin: 0 auto;
            background-color: #FAF8F5;
            padding: 30px;
            border-radius: 8px;
        }
        
        h1 {
            color: #FF8C42;
            margin-bottom: 10px;
        }
        
        .date-range {
            color: #6B6B6B;
            font-size: 14px;
            margin-bottom: 30px;
        }
        
        .section {
            margin-bottom: 30px;
        }
        
        .section-title {
            color: #2C2C2C;
            font-size: 18px;
            font-weight: 600;
            margin-bottom: 15px;
            border-bottom: 2px solid #E8DCC8;
            padding-bottom: 5px;
        }
        
        .stat-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }
        
        .stat-card {
            background-color: #FFFFFF;
            padding: 15px;
            border-radius: 5px;
            border: 1px solid #E8DCC8;
        }
        
        .stat-label {
            font-size: 12px;
            color: #6B6B6B;
            margin-bottom: 5px;
        }
        
        .stat-value {
            font-size: 20px;
            font-weight: 600;
            color: #FF8C42;
        }
        
        .insight {
            background-color: #FFFFFF;
            padding: 15px;
            border-left: 4px solid #FF8C42;
            margin-bottom: 15px;
        }
        
        .insight-title {
            font-weight: 600;
            margin-bottom: 8px;
        }
        
        .insight-text {
            font-size: 14px;
            color: #4B4B4B;
        }
        
        .chart {
            text-align: center;
            margin: 20px 0;
        }
        
        .chart img {
            max-width: 100%;
            height: auto;
            border-radius: 5px;
        }
        
        .footer {
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #E8DCC8;
            font-size: 12px;
            color: #6B6B6B;
            text-align: center;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>{{ title }}</h1>
        <div class="date-range">
            {{ date_range[0].strftime('%B %d, %Y') }} - {{ date_range[1].strftime('%B %d, %Y') }}
        </div>
        
        {% if summary_stats %}
        <div class="section">
            <h2 class="section-title">Summary Statistics</h2>
            <div class="stat-grid">
                {% for metric, stats in summary_stats.items() | list[:4] %}
                <div class="stat-card">
                    <div class="stat-label">{{ metric.replace('_', ' ').title() }}</div>
                    <div class="stat-value">{{ "%.1f"|format(stats.mean) }}</div>
                </div>
                {% endfor %}
            </div>
        </div>
        {% endif %}
        
        {% if insights %}
        <div class="section">
            <h2 class="section-title">Key Insights</h2>
            {% for insight in insights[:3] %}
            <div class="insight">
                <div class="insight-title">{{ insight.title }}</div>
                <div class="insight-text">{{ insight.summary }}</div>
            </div>
            {% endfor %}
        </div>
        {% endif %}
        
        {% if charts %}
        <div class="section">
            <h2 class="section-title">Visualizations</h2>
            {% for chart in charts %}
            <div class="chart">
                <img src="data:image/png;base64,{{ chart.data }}" alt="{{ chart.title }}">
                <p>{{ chart.title }}</p>
            </div>
            {% endfor %}
        </div>
        {% endif %}
        
        <div class="footer">
            <p>Generated by Apple Health Monitor on {{ generated_date.strftime('%B %d, %Y at %I:%M %p') }}</p>
            <p>This report is for informational purposes only.</p>
        </div>
    </div>
</body>
</html>
        """)
        
        return template.render(**data)
        
    def _create_comprehensive_template(self) -> Template:
        """Create comprehensive report template."""
        # This is simplified - in production, load from file
        return Template(self._get_comprehensive_template_content())
        
    def _create_medical_template(self) -> Template:
        """Create medical summary template."""
        return Template(self._get_medical_template_content())
        
    def _create_progress_template(self) -> Template:
        """Create progress tracking template."""
        return Template(self._get_progress_template_content())
        
    def _create_weekly_template(self) -> Template:
        """Create weekly summary template."""
        return Template(self._get_weekly_template_content())
        
    def _create_monthly_template(self) -> Template:
        """Create monthly summary template."""
        return Template(self._get_monthly_template_content())
        
    def _get_comprehensive_template_content(self) -> str:
        """Get comprehensive template content."""
        # Simplified version - in production, this would be more elaborate
        return """
        <h1>{{ title or "Comprehensive Health Report" }}</h1>
        <p>{{ date_range[0] }} to {{ date_range[1] }}</p>
        
        {% if insights %}
        <h2>Executive Summary</h2>
        <p><strong>Key Finding:</strong> {{ insights[0].title }}</p>
        <p>{{ insights[0].summary }}</p>
        {% endif %}
        
        <h2>Health Metrics</h2>
        {% for metric, stats in summary_stats.items() %}
        <h3>{{ metric }}</h3>
        <ul>
            <li>Average: {{ stats.mean }}</li>
            <li>Range: {{ stats.min }} - {{ stats.max }}</li>
            <li>Trend: {{ stats.trend }}</li>
        </ul>
        {% endfor %}
        """
        
    def _get_medical_template_content(self) -> str:
        """Get medical summary template content."""
        return """
        <h1>Medical Summary Report</h1>
        <p>Prepared for healthcare provider review</p>
        
        <h2>Patient Health Metrics</h2>
        {% for metric, stats in summary_stats.items() %}
        <div>
            <h3>{{ metric }}</h3>
            <table>
                <tr><td>Mean:</td><td>{{ stats.mean }}</td></tr>
                <tr><td>Min:</td><td>{{ stats.min }}</td></tr>
                <tr><td>Max:</td><td>{{ stats.max }}</td></tr>
            </table>
        </div>
        {% endfor %}
        
        <p><em>This report is for informational purposes only.</em></p>
        """
        
    def _get_progress_template_content(self) -> str:
        """Get progress tracking template content."""
        return """
        <h1>Progress Report</h1>
        
        <h2>Goal Progress</h2>
        {% for metric, stats in summary_stats.items() %}
        <div>
            <h3>{{ metric }}</h3>
            <p>Current: {{ stats.mean }} | Trend: {{ stats.trend }}</p>
        </div>
        {% endfor %}
        """
        
    def _get_weekly_template_content(self) -> str:
        """Get weekly summary template content."""
        return """
        <h1>Weekly Health Summary</h1>
        <p>Week of {{ date_range[0].strftime('%B %d, %Y') }}</p>
        
        <h2>This Week's Highlights</h2>
        {% if insights %}
        {% for insight in insights[:3] %}
        <p>• {{ insight.title }}</p>
        {% endfor %}
        {% endif %}
        """
        
    def _get_monthly_template_content(self) -> str:
        """Get monthly summary template content."""
        return """
        <h1>Monthly Health Report</h1>
        <p>{{ date_range[0].strftime('%B %Y') }}</p>
        
        <h2>Monthly Overview</h2>
        {% for metric, stats in summary_stats.items() %}
        <p>{{ metric }}: {{ stats.mean }} ({{ stats.trend }})</p>
        {% endfor %}
        """